import glob
import pandas as pd
import openpyxl # Need to read xlsx
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import PatternFill

import read_settings

def convert_color_name(color):
    """
    Converts a color name to its RGB value.
    Args:
        color (str): The name of the color.
    Returns:
        tuple: The Color code (e.g., 'FF0000' for red).
    Example:
        >>> convert_color_name("red")
        'FF0000'
        >>> convert_color_name("unknown")
        'FFFF00'  # default to yellow
    """
    # if color matches "#xxxxxx" format
    if color is None:
        return 'FFFF00'
    if color[0] == "#":
        return color[1:]
    COLORS = {
        "white" : 'FFFFFF',
        "purple": 'FF00FF',
        "yellow": 'FFFF00',
        "red"   : 'FF0000',
        "sky"   : '00FFFF',
        "blue"  : '0000FF',
        "green" : '00FF00',
        "gray"  : 'cccccc' 
    }
    try:
        col = COLORS[color]
    except:
        col = 'FFFF00'
    return col


def highlight_xlsx(xlsx, keywords="", colors="yellow", exact=False):
    out_xlsx = xlsx.name.replace(".xlsx", "_highlighted.xlsx")
    wb = openpyxl.load_workbook(xlsx)
    sheets = wb.sheetnames
    setting_sheet = "setting_for_highlight"
    if setting_sheet in sheets:
        colnames = ["keywords", "colors"]
        keywords, colors = read_settings.read_settings(wb, sheet=setting_sheet, colnames=colnames)
    if isinstance(keywords, str) and keywords == "":
        wb.save(out_xlsx) # no change
        return out_xlsx
    offset = 64 # need to convert number to character
    for ws in wb.worksheets:
        max_row = ws.max_row
        max_col = ws.max_column
        range_str = "".join([chr(1 + offset), str(1), ":", chr(max_col + offset), str(max_row)])
        if isinstance(keywords, str):
            keywords = keywords.split(";")
        # if color is a single color
        if isinstance(colors, str):
            colors = [colors] * len(keywords)
        for kwd, clr in zip(keywords, colors):
            highlight_cell(ws, range_str, kwd, convert_color_name(clr), exact=exact)
    wb.save(out_xlsx)
    return out_xlsx

def highlight_cell(sheet, range_str, keyword, color, exact=False):
    if keyword is None:
        pass
    else:
        color_fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
        if exact:
            condition = f'EXACT("{keyword}", A1)'
        else:
            condition = f'SEARCH("{keyword}", A1)'
        rule = FormulaRule(formula=[condition], fill=color_fill)
        sheet.conditional_formatting.add(range_str, rule)

if __name__ == "__main__":
    import os

    input_xlsxs = glob.glob("*.xlsx")

    for xlsx in input_xlsxs:
        if not "highlighted" in xlsx:
            highlighted = highlight_xlsx(xlsx)
            os.startfile(highlighted)
